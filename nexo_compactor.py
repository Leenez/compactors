from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('nexo_transactions.xlsx')
wb2 = Workbook()
ws = wb.active
ws2 = wb2.active
first_row = ["Transaction","Type","Currency","Amount","USD Equivalent","Details","Outstanding Loan","Date / Time"]
ws2.append(first_row)

nexointerest = {
    "amount": 0,
    "usdeq": float(0.0),
    "date": 0
}

for row in range(2, ws.max_row):
    if (ws["B" + str(row)].value == "Interest" or ws["B" + str(row)].value == "FixedTermInterest") and ws["C" + str(row)].value == "NEXONEXO":
        nexointerest["date"] = ws["H" + str(row)].value
        break

for row in range(2, ws.max_row):
    transaction = ws["A" + str(row)].value
    type_ = ws["B" + str(row)].value
    currency =  ws["C" + str(row)].value
    amount = ws["D" + str(row)].value
    usdeq = float(ws["E" + str(row)].value[1:])
    details = ws["F" + str(row)].value
    outstanding = ws["G" + str(row)].value
    date_ = ws["H" + str(row)].value

    if (type_ == "Interest" or type_ == "FixedTermInterest") and currency == "NEXONEXO":
        if nexointerest["date"].split(" ")[0] == date_.split(" ")[0]:
            nexointerest["amount"] = nexointerest["amount"] + amount
            nexointerest["usdeq"] = float(nexointerest["usdeq"]) + float(usdeq)
        else:
            new_row = [transaction,type_,currency,nexointerest["amount"],"$" + str(nexointerest["usdeq"]),details,outstanding,nexointerest["date"]]
            ws2.append(new_row)
            nexointerest["date"] = date_
            nexointerest["amount"] = amount
            nexointerest["usdeq"] = usdeq
    else:
        data = [transaction,type_,currency,amount,"$" + str(usdeq),details,outstanding,date_]
        ws2.append(data)

if (type_ == "Interest" or type_ == "FixedTermInterest") and currency == "NEXONEXO":
    new_row = [transaction,type_,currency,nexointerest["amount"],"$" + str(nexointerest["usdeq"]),details,outstanding,nexointerest["date"]]
    ws2.append(new_row)
    
wb2.save('nexo_compact_data.xlsx')
        
    

#print(ws["A" + str(row)].value)
    #for col in range(1, ws.max_column):
        #char = get_column_letter(col)
        #print(ws[char + str(row)].value)
#ws.append(['something1', 'something2'])
#ws.insert_cols(2)
#ws.delete_cols(3)
#ws['A1'].value = "Asset"
#print(ws.max_row)
#print(ws.max_column)
#print(wb.sheetnames)
#print(wb['Sheet1'])
#print(ws['A1'].value)
#ws.merge_cells("A1:A2")
#headings = ['Name'] + data['Joe'].keys()
#ws.append(headings)
#for col in range(2, len(data['Joe'] + 2))
#wb.save('holdings.xlsx')

# if row + 1 == ws.max_row:
#         data_hpm = [ws["A" + str(row - 1)].value , ws["B" + str(row - 1)].value , "Hashpower mining" , hashpowermining["amountbtc"] , hashpowermining["rate"] / hashpowermining["counter"] , hashpowermining["amounteur"]]
#         ws2.append(data_hpm)
#         data_hpmf = [ws["A" + str(row )].value , ws["B" + str(row)].value , "Hashpower mining fee" , hashpowerminingfee["amountbtc"] , hashpowerminingfee["rate"] / hashpowerminingfee["counter"] , hashpowerminingfee["amounteur"]]
#         ws2.append(data_hpmf)     
#     elif ws["C" + str(row)].value == "Hashpower mining":
#         hashpowermining["amountbtc"] = hashpowermining["amountbtc"] + ws["D" + str(row)].value
#         hashpowermining["rate"] = hashpowermining["rate"] + ws["E" + str(row)].value
#         hashpowermining["amounteur"] = hashpowermining["amounteur"] + ws["F" + str(row)].value
#         hashpowermining["counter"] = hashpowermining["counter"] + 1
#     elif ws["C" + str(row)].value == "Hashpower mining fee":
#         hashpowerminingfee["amountbtc"] = hashpowerminingfee["amountbtc"] + ws["D" + str(row)].value
#         hashpowerminingfee["rate"] = hashpowerminingfee["rate"] + ws["E" + str(row)].value
#         hashpowerminingfee["amounteur"] = hashpowerminingfee["amounteur"] + ws["F" + str(row)].value
#         hashpowerminingfee["counter"] = hashpowerminingfee["counter"] + 1
#     elif ws["C" + str(row)].value == "Withdrawal complete":        
#         data_hpm = [ws["A" + str(row - 2)].value , ws["B" + str(row - 2)].value , "Hashpower mining" , hashpowermining["amountbtc"] , hashpowermining["rate"] / hashpowermining["counter"] , hashpowermining["amounteur"]]
#         ws2.append(data_hpm)
#         data_hpmf = [ws["A" + str(row - 1)].value , ws["B" + str(row - 1)].value , "Hashpower mining fee" , hashpowerminingfee["amountbtc"] , hashpowerminingfee["rate"] / hashpowerminingfee["counter"] , hashpowerminingfee["amounteur"]]
#         ws2.append(data_hpmf)
#         withdrawal = [ws["A" + str(row )].value, ws["B" + str(row )].value , ws["C" + str(row )].value , ws["D" + str(row )].value , ws["E" + str(row )].value , ws["F" + str(row )].value]
#         ws2.append(withdrawal)
#         hashpowermining["amountbtc"] = 0
#         hashpowermining["rate"] = 0
#         hashpowermining["amounteur"] = 0
#         hashpowermining["counter"] = 0
#         hashpowerminingfee["amountbtc"] = 0
#         hashpowerminingfee["rate"] = 0
#         hashpowerminingfee["amounteur"] = 0
#         hashpowerminingfee["counter"] = 0
#     else:
#         print("SHIT")
#         break